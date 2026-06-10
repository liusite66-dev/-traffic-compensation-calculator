"""
交通事故赔偿金额计算器
基于 CustomTkinter 的桌面应用
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# ── 全局设置 ──
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Google 风格配色 ──
COLORS = {
    "header_bg": "#1a73e8",       # Google Blue
    "header_fg": "#ffffff",
    "card_bg": "#f8f9fa",         # 浅灰卡片
    "card_border": "#e8eaed",     # 边框灰
    "accent": "#1a73e8",
    "text_primary": "#202124",
    "text_secondary": "#5f6368",
    "success_bg": "#e6f4ea",
    "success_fg": "#1e7e34",
    "font": "Microsoft YaHei",
}

# ── 项目定义 ──
MEDICAL_ITEMS = [
    "医疗费",
    "住院伙食补助费",
    "营养费",
]

DEATH_DISABILITY_ITEMS = [
    "护理费",
    "误工费",
    "残疾赔偿金",
    "被扶养人生活费",
    "精神损害抚慰金",
    "交通费",
    "护理用品费",
]

PROPERTY_ITEMS = [
    "财产损失",
]

ALL_ITEMS = MEDICAL_ITEMS + DEATH_DISABILITY_ITEMS + PROPERTY_ITEMS

CATEGORY_NAMES = {
    "medical": "医疗费类",
    "death_disability": "死亡伤残赔偿金类",
    "property": "财产类",
}

# 交强险限额
JIAO_QIANG_LIMITS = {
    "medical": 18000,
    "death_disability": 180000,
    "property": 2000,
}


class CompensationApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("交通事故赔偿金额计算器")
        self.geometry("1200x800")
        self.minsize(1000, 650)

        # ── 原告信息 ──
        self.plaintiff_name = ""

        # ── 数据存储 ──
        # items[item_name] = {"claim": float, "support": float}
        self.items = {}
        for item in ALL_ITEMS:
            self.items[item] = {"claim": 0.0, "support": 0.0}

        # 额外参数
        self.extra = {
            "损失不分责": 0.0,
            "被告支付": 0.0,
            "总计支持": 0.0,
            "案件受理费": 0.0,
            "被告承担受理费": 0.0,
            "鉴定费": 0.0,
            "被告垫付总额": 0.0,
            "原告三者险应得": 0.0,
        }

        # 交强险扣除 (每个板块)
        self.jiao_qiang_deduct = {
            "medical": 0.0,
            "death_disability": 0.0,
            "property": 0.0,
        }

        # 承担比例 (如 70% -> 0.7)
        self.responsibility_ratio = 1.0

        # 医疗费逐笔明细
        self.medical_detail_rows = []  # list of {"date_var": StringVar, "amount_var": StringVar, "frame": CTkFrame}

        # 防抖计时器
        self._calc_after_id = None

        # ── UI ──
        self._build_ui()

    # ══════════════════════════════════════════
    # UI 构建
    # ══════════════════════════════════════════

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        main_frame = ctk.CTkFrame(self, fg_color=COLORS["card_bg"])
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(2, weight=1)

        # ── 顶部导航栏 ──
        header = ctk.CTkFrame(main_frame, fg_color=COLORS["header_bg"], height=56, corner_radius=0)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(2, weight=1)
        header.grid_propagate(False)

        ctk.CTkLabel(header, text="交通事故赔偿计算器",
                      font=(COLORS["font"], 18, "bold"),
                      text_color=COLORS["header_fg"]).grid(row=0, column=0, padx=(24, 10), pady=12, sticky="w")

        # 原告姓名（导航栏内联）
        ctk.CTkLabel(header, text="原告", font=(COLORS["font"], 13, "bold"),
                      text_color=COLORS["header_fg"]).grid(row=0, column=1, padx=(0, 4), sticky="e")
        self.plaintiff_var = tk.StringVar()
        plaintiff_entry = ctk.CTkEntry(header, width=180, height=32,
                                        font=(COLORS["font"], 13, "bold"),
                                        textvariable=self.plaintiff_var,
                                        fg_color="#ffffff", text_color="#202124",
                                        border_width=0, corner_radius=4)
        plaintiff_entry.grid(row=0, column=2, padx=(0, 24), pady=10, sticky="w")
        self.plaintiff_var.trace_add("write", lambda *args: setattr(self, 'plaintiff_name', self.plaintiff_var.get().strip()))

        # ── 主体区域 ──
        self.tabview = ctk.CTkTabview(main_frame, fg_color="transparent",
                                       segmented_button_fg_color=COLORS["card_border"],
                                       segmented_button_selected_color=COLORS["accent"],
                                       segmented_button_selected_hover_color="#1557b0",
                                       segmented_button_unselected_color=COLORS["card_bg"],
                                       segmented_button_unselected_hover_color="#e8eaed",
                                       corner_radius=8)
        # 统一设置标签文字为深色，确保清晰可读
        try:
            self.tabview._segmented_button.configure(text_color="#202124",
                                                      font=(COLORS["font"], 13, "bold"))
        except Exception:
            pass
        self.tabview.grid(row=2, column=0, sticky="nsew", padx=16, pady=(8, 16))

        # 创建各个 Tab
        self.tab_medical = self.tabview.add("医疗费类")
        self.tab_death = self.tabview.add("死亡伤残类")
        self.tab_property = self.tabview.add("财产类")
        self.tab_summary = self.tabview.add("汇总计算")
        self.tab_other = self.tabview.add("其他费用")
        self.tab_export = self.tabview.add("导出Excel")

        self._build_medical_tab()
        self._build_death_tab()
        self._build_property_tab()
        self._build_summary_tab()
        self._build_other_tab()
        self._build_export_tab()

    # ── 卡片区 ──

    def _make_card(self, parent, title, start_row, colspan=3):
        """创建卡片容器，返回 card_frame 和 内部 frame"""
        sep = ctk.CTkLabel(parent, text="", height=6, fg_color="transparent")
        sep.grid(row=start_row, column=0, columnspan=colspan)
        title_lbl = ctk.CTkLabel(parent, text=title, font=(COLORS["font"], 13, "bold"),
                                  text_color=COLORS["text_primary"], anchor="w")
        title_lbl.grid(row=start_row + 1, column=0, columnspan=colspan, padx=4, pady=(4, 0), sticky="w")
        card = ctk.CTkFrame(parent, fg_color=COLORS["card_bg"], corner_radius=8,
                             border_width=1, border_color=COLORS["card_border"])
        card.grid(row=start_row + 2, column=0, columnspan=colspan, sticky="ew", padx=2, pady=(4, 8))
        for i in range(colspan):
            card.grid_columnconfigure(i, weight=0)
        return card

    # ── 通用: 创建输入行 ──

    def _make_item_row(self, parent, row, item_name, claim_var, support_var, card=None):
        """创建一行: 项目名 | 原告要求输入 | 本院支持输入"""
        bg = "transparent" if card else None
        label = ctk.CTkLabel(parent, text=item_name, width=120, anchor="w",
                              font=(COLORS["font"], 13), text_color=COLORS["text_primary"])
        label.grid(row=row, column=0, padx=(14, 8), pady=4, sticky="w")

        claim_entry = ctk.CTkEntry(parent, width=130, height=34,
                                    font=(COLORS["font"], 13),
                                    textvariable=claim_var,
                                    border_width=1, border_color=COLORS["card_border"],
                                    corner_radius=6)
        claim_entry.grid(row=row, column=1, padx=4, pady=4)

        support_entry = ctk.CTkEntry(parent, width=130, height=34,
                                      font=(COLORS["font"], 13),
                                      textvariable=support_var,
                                      border_width=1, border_color=COLORS["card_border"],
                                      corner_radius=6)
        support_entry.grid(row=row, column=2, padx=4, pady=4)

    def _make_header_row(self, parent, row, col_start=0):
        """创建表头 - 与数据行对齐"""
        headers = ["项目", "原告要求", "本院支持"]
        widths = [120, 130, 130]
        paddings = [(14, 8), 4, 4]
        for i, h in enumerate(headers):
            lbl = ctk.CTkLabel(parent, text=h, font=(COLORS["font"], 12, "bold"),
                                text_color=COLORS["text_secondary"], width=widths[i])
            p = paddings[i]
            kwargs = {"row": row, "column": col_start + i, "padx": p, "pady": (4, 2)}
            if i == 0:
                kwargs["sticky"] = "w"
            lbl.grid(**kwargs)

    def _make_subtotal_row(self, parent, row, claim_var, support_var):
        """小计行"""
        ctk.CTkLabel(parent, text="小计", font=(COLORS["font"], 13, "bold"),
                      text_color=COLORS["accent"]).grid(row=row, column=0, padx=(10, 5), pady=5)
        claim_entry = ctk.CTkEntry(parent, width=130, height=34,
                                    font=(COLORS["font"], 13, "bold"),
                                    textvariable=claim_var, state="readonly",
                                    fg_color=COLORS["success_bg"], text_color=COLORS["success_fg"],
                                    border_width=0, corner_radius=6)
        claim_entry.grid(row=row, column=1, padx=5, pady=5)
        support_entry = ctk.CTkEntry(parent, width=130, height=34,
                                      font=(COLORS["font"], 13, "bold"),
                                      textvariable=support_var, state="readonly",
                                      fg_color=COLORS["success_bg"], text_color=COLORS["success_fg"],
                                      border_width=0, corner_radius=6)
        support_entry.grid(row=row, column=2, padx=5, pady=5)

    # ── Tab 1: 医疗费类 ──

    def _build_medical_tab(self):
        self.tab_medical.grid_columnconfigure(0, weight=1)
        self.tab_medical.grid_rowconfigure(0, weight=1)

        frame = ctk.CTkScrollableFrame(self.tab_medical)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        frame.grid_columnconfigure(0, weight=0)
        frame.grid_columnconfigure(1, weight=0)
        frame.grid_columnconfigure(2, weight=0)

        self._make_header_row(frame, 0)

        self.medical_vars = {}
        for i, item in enumerate(["医疗费"]):
            claim_var = tk.StringVar(value="0")
            support_var = tk.StringVar(value="0")
            self.medical_vars[item] = {"claim": claim_var, "support": support_var}
            self._make_item_row(frame, i + 1, item, claim_var, support_var)
            claim_var.trace_add("write", lambda *args: self._schedule_calc())
            support_var.trace_add("write", lambda *args: self._schedule_calc())

        # ── 医疗费逐笔明细 ──
        detail_label_row = 2
        ctk.CTkLabel(frame, text="▼ 医疗费逐笔明细", font=("Microsoft YaHei", 12, "bold")).grid(
            row=detail_label_row, column=0, columnspan=3, pady=(10, 2), sticky="w")

        detail_header_row = 3
        ctk.CTkLabel(frame, text="日期（选填）", font=("Microsoft YaHei", 11)).grid(row=detail_header_row, column=0, padx=5, pady=1)
        ctk.CTkLabel(frame, text="金额（元）", font=("Microsoft YaHei", 11)).grid(row=detail_header_row, column=1, padx=5, pady=1)

        # 容器 frame — 所有明细行用 pack 管理，不影响主 grid
        self.medical_detail_container = ctk.CTkFrame(frame, fg_color="transparent")
        self.medical_detail_container.grid(row=4, column=0, columnspan=3, sticky="ew", padx=5)

        # 添加行按钮 + 合计
        detail_btn_row = 5
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=detail_btn_row, column=0, columnspan=3, pady=5)

        add_btn = ctk.CTkButton(btn_frame, text="+ 添加一行", command=self._add_medical_detail_row,
                                 width=100, height=28)
        add_btn.pack(side="left", padx=5)

        self.medical_detail_total_var = tk.StringVar(value="逐笔合计: 0.00 元")
        ctk.CTkLabel(btn_frame, textvariable=self.medical_detail_total_var,
                      font=("Microsoft YaHei", 12, "bold")).pack(side="left", padx=15)

        # ── 住院伙食补助费 ──
        row7 = 7
        self.hospital_claim_var = tk.StringVar(value="0")
        self.hospital_support_var = tk.StringVar(value="0")
        self.medical_vars["住院伙食补助费"] = {"claim": self.hospital_claim_var, "support": self.hospital_support_var}
        self._make_item_row(frame, row7, "住院伙食补助费", self.hospital_claim_var, self.hospital_support_var)
        self.hospital_claim_var.trace_add("write", lambda *args: self._schedule_calc())
        self.hospital_support_var.trace_add("write", lambda *args: self._schedule_calc())

        # ── 营养费（天数 × 50 自动计算）──
        row8 = 8
        nutrition_claim_var = tk.StringVar(value="0")
        nutrition_support_var = tk.StringVar(value="0")
        self.medical_vars["营养费"] = {"claim": nutrition_claim_var, "support": nutrition_support_var}

        # 营养费 — 第一行：项目名 + 原告要求 + 本院支持（自动/只读）
        ctk.CTkLabel(frame, text="营养费", width=120, anchor="w", font=("Microsoft YaHei", 13)).grid(
            row=row8, column=0, padx=(10, 5), pady=(8, 0), sticky="w")
        ctk.CTkEntry(frame, width=130, textvariable=nutrition_claim_var).grid(
            row=row8, column=1, padx=5, pady=(8, 0))
        self.nutrition_support_entry = ctk.CTkEntry(frame, width=130, textvariable=nutrition_support_var, state="readonly")
        self.nutrition_support_entry.grid(row=row8, column=2, padx=5, pady=(8, 0))
        nutrition_claim_var.trace_add("write", lambda *args: self._schedule_calc())
        # support_var 是自动计算的，不需要 trace

        # 营养费 — 第二行：天数输入 + 计算
        row9 = 9
        ctk.CTkLabel(frame, text="  ﹂ 天数:", font=("Microsoft YaHei", 12)).grid(
            row=row9, column=0, padx=(10, 5), pady=1, sticky="w")
        self.nutrition_days_var = tk.StringVar(value="0")
        days_entry = ctk.CTkEntry(frame, width=100, textvariable=self.nutrition_days_var)
        days_entry.grid(row=row9, column=1, padx=5, pady=1, sticky="w")
        ctk.CTkLabel(frame, text="× 50 元/天", font=("Microsoft YaHei", 12)).grid(
            row=row9, column=1, padx=40, pady=1, sticky="w")
        self.nutrition_calc_label = tk.StringVar(value="= 0.00 元")
        ctk.CTkLabel(frame, textvariable=self.nutrition_calc_label, font=("Microsoft YaHei", 12, "bold")).grid(
            row=row9, column=2, padx=5, pady=1, sticky="w")
        self.nutrition_days_var.trace_add("write", lambda *args: self._update_nutrition())

        # 营养费 — 第三行：依据选择（单选）
        row10 = 10
        radio_frame = ctk.CTkFrame(frame, fg_color="transparent")
        radio_frame.grid(row=row10, column=0, columnspan=3, pady=(2, 5), sticky="w")
        ctk.CTkLabel(radio_frame, text="  ﹂ 依据:", font=("Microsoft YaHei", 12)).pack(side="left")
        self.nutrition_basis_var = tk.StringVar(value="医嘱/鉴定")
        ctk.CTkRadioButton(radio_frame, text="按医嘱/鉴定确定营养期",
                           variable=self.nutrition_basis_var, value="医嘱/鉴定",
                           font=("Microsoft YaHei", 12)).pack(side="left", padx=5)
        ctk.CTkRadioButton(radio_frame, text="按三期鉴定确定营养期",
                           variable=self.nutrition_basis_var, value="三期鉴定",
                           font=("Microsoft YaHei", 12)).pack(side="left", padx=5)

        # 小计
        subtotal_row = 12
        self.medical_claim_total = tk.StringVar(value="0.00")
        self.medical_support_total = tk.StringVar(value="0.00")
        self._make_subtotal_row(frame, subtotal_row, self.medical_claim_total, self.medical_support_total)

        # 交强险
        row = subtotal_row + 1
        sep = ctk.CTkLabel(frame, text="─" * 40)
        sep.grid(row=row, column=0, columnspan=3, pady=5)

        row += 1
        ctk.CTkLabel(frame, text="交强险医疗费限额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        ctk.CTkLabel(frame, text=f"{JIAO_QIANG_LIMITS['medical']} 元", font=("Microsoft YaHei", 12)).grid(row=row, column=1, padx=5, pady=3, sticky="w")

        row += 1
        ctk.CTkLabel(frame, text="交强险扣除金额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        self.jq_medical_var = tk.StringVar(value="0")
        jq_entry = ctk.CTkEntry(frame, width=130, textvariable=self.jq_medical_var)
        jq_entry.grid(row=row, column=1, padx=5, pady=3)
        self.jq_medical_var.trace_add("write", lambda *args: self._schedule_calc())

        # 预置一行
        self._add_medical_detail_row()

    # ── 医疗费逐笔明细 ──

    def _add_medical_detail_row(self, date="", amount=""):
        """添加一行医疗费明细"""
        row_frame = ctk.CTkFrame(self.medical_detail_container)
        row_frame.pack(fill="x", pady=1)

        date_var = tk.StringVar(value=date)
        amount_var = tk.StringVar(value=amount)

        date_entry = ctk.CTkEntry(row_frame, width=130, textvariable=date_var)
        date_entry.pack(side="left", padx=2)

        amount_entry = ctk.CTkEntry(row_frame, width=130, textvariable=amount_var)
        amount_entry.pack(side="left", padx=2)

        del_btn = ctk.CTkButton(row_frame, text="×", width=30, height=28,
                                 fg_color="#e53935", hover_color="#c62828")
        del_btn.pack(side="left", padx=2)

        # 用闭包绑定删除功能
        def on_delete():
            self._remove_medical_detail_row(row_frame)

        del_btn.configure(command=on_delete)

        # 金额变更时更新合计
        def on_amount_change(*args):
            self._update_medical_detail_total()

        amount_var.trace_add("write", on_amount_change)
        date_var.trace_add("write", on_amount_change)

        self.medical_detail_rows.append({
            "date_var": date_var,
            "amount_var": amount_var,
            "frame": row_frame,
        })

        self._update_medical_detail_total()

    def _remove_medical_detail_row(self, row_frame):
        """删除一行医疗费明细"""
        for i, row_data in enumerate(self.medical_detail_rows):
            if row_data["frame"] == row_frame:
                row_data["frame"].destroy()
                del self.medical_detail_rows[i]
                break
        self._update_medical_detail_total()

    def _update_nutrition(self):
        """计算营养费 = 天数 × 50"""
        try:
            days = float(self.nutrition_days_var.get().strip() or "0")
        except ValueError:
            days = 0.0
        total = days * 50
        self.nutrition_calc_label.set(f"= {total:.2f} 元")
        self.medical_vars["营养费"]["support"].set(f"{total:.2f}")
        self._schedule_calc()

    def _update_medical_detail_total(self):
        """更新医疗费明细合计，并同步到医疗费 本院支持"""
        total = 0.0
        for row_data in self.medical_detail_rows:
            try:
                v = row_data["amount_var"].get().strip()
                if v:
                    total += float(v)
            except (ValueError, tk.TclError):
                pass

        self.medical_detail_total_var.set(f"逐笔合计: {total:.2f} 元")
        # 同步到医疗费本院支持
        self.medical_vars["医疗费"]["support"].set(f"{total:.2f}")
        self._schedule_calc()

    # ── 获取明细数据供导出 ──

    def _get_medical_details(self):
        """返回 [(日期, 金额), ...] 列表"""
        result = []
        for row_data in self.medical_detail_rows:
            date = row_data["date_var"].get().strip()
            amount_str = row_data["amount_var"].get().strip()
            try:
                amount = float(amount_str) if amount_str else 0.0
            except ValueError:
                amount = 0.0
            result.append((date, amount))
        return result

    # ── Tab 2: 死亡伤残类 ──

    def _build_death_tab(self):
        self.tab_death.grid_columnconfigure(0, weight=0)
        self.tab_death.grid_columnconfigure(1, weight=0)
        self.tab_death.grid_columnconfigure(2, weight=0)

        frame = ctk.CTkScrollableFrame(self.tab_death)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        frame.grid_columnconfigure(0, weight=0)
        frame.grid_columnconfigure(1, weight=0)
        frame.grid_columnconfigure(2, weight=0)

        self._make_header_row(frame, 0)

        # ── 卡片: 死亡伤残赔偿项目 ──
        card = ctk.CTkFrame(frame, fg_color=COLORS["card_bg"], corner_radius=8,
                             border_width=1, border_color=COLORS["card_border"])
        card.grid(row=1, column=0, columnspan=3, sticky="ew", padx=2, pady=(4, 8))
        card.grid_columnconfigure(0, weight=0)
        card.grid_columnconfigure(1, weight=0)
        card.grid_columnconfigure(2, weight=0)

        self.death_vars = {}
        for i, item in enumerate(DEATH_DISABILITY_ITEMS):
            claim_var = tk.StringVar(value="0")
            support_var = tk.StringVar(value="0")
            self.death_vars[item] = {"claim": claim_var, "support": support_var}
            self._make_item_row(card, i, item, claim_var, support_var)
            claim_var.trace_add("write", lambda *args: self._schedule_calc())
            support_var.trace_add("write", lambda *args: self._schedule_calc())

        # 小计
        row = 2
        self.death_claim_total = tk.StringVar(value="0.00")
        self.death_support_total = tk.StringVar(value="0.00")
        self._make_subtotal_row(frame, row, self.death_claim_total, self.death_support_total)

        # 交强险
        row += 1
        sep = ctk.CTkLabel(frame, text="─" * 40)
        sep.grid(row=row, column=0, columnspan=3, pady=5)

        row += 1
        ctk.CTkLabel(frame, text="交强险死亡伤残限额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        ctk.CTkLabel(frame, text=f"{JIAO_QIANG_LIMITS['death_disability']} 元", font=("Microsoft YaHei", 12)).grid(row=row, column=1, padx=5, pady=3, sticky="w")

        row += 1
        ctk.CTkLabel(frame, text="交强险扣除金额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        self.jq_death_var = tk.StringVar(value="0")
        jq_entry = ctk.CTkEntry(frame, width=130, textvariable=self.jq_death_var)
        jq_entry.grid(row=row, column=1, padx=5, pady=3)
        self.jq_death_var.trace_add("write", lambda *args: self._schedule_calc())

    # ── Tab 3: 财产类 ──

    def _build_property_tab(self):
        self.tab_property.grid_columnconfigure(0, weight=0)
        self.tab_property.grid_columnconfigure(1, weight=0)
        self.tab_property.grid_columnconfigure(2, weight=0)

        frame = ctk.CTkFrame(self.tab_property)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        frame.grid_columnconfigure(0, weight=0)
        frame.grid_columnconfigure(1, weight=0)
        frame.grid_columnconfigure(2, weight=0)

        self._make_header_row(frame, 0)

        # ── 卡片: 财产损失项目 ──
        card = ctk.CTkFrame(frame, fg_color=COLORS["card_bg"], corner_radius=8,
                             border_width=1, border_color=COLORS["card_border"])
        card.grid(row=1, column=0, columnspan=3, sticky="ew", padx=2, pady=(4, 8))
        card.grid_columnconfigure(0, weight=0)
        card.grid_columnconfigure(1, weight=0)
        card.grid_columnconfigure(2, weight=0)

        self.property_vars = {}
        for i, item in enumerate(PROPERTY_ITEMS):
            claim_var = tk.StringVar(value="0")
            support_var = tk.StringVar(value="0")
            self.property_vars[item] = {"claim": claim_var, "support": support_var}
            self._make_item_row(card, i, item, claim_var, support_var)
            claim_var.trace_add("write", lambda *args: self._schedule_calc())
            support_var.trace_add("write", lambda *args: self._schedule_calc())

        # 小计
        row = 2
        self.property_claim_total = tk.StringVar(value="0.00")
        self.property_support_total = tk.StringVar(value="0.00")
        self._make_subtotal_row(frame, row, self.property_claim_total, self.property_support_total)

        # 交强险
        row += 1
        sep = ctk.CTkLabel(frame, text="─" * 40)
        sep.grid(row=row, column=0, columnspan=3, pady=5)

        row += 1
        ctk.CTkLabel(frame, text="交强险财产损失限额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        ctk.CTkLabel(frame, text=f"{JIAO_QIANG_LIMITS['property']} 元", font=("Microsoft YaHei", 12)).grid(row=row, column=1, padx=5, pady=3, sticky="w")

        row += 1
        ctk.CTkLabel(frame, text="交强险扣除金额:", font=("Microsoft YaHei", 12)).grid(row=row, column=0, padx=(10, 5), pady=3, sticky="w")
        self.jq_property_var = tk.StringVar(value="0")
        jq_entry = ctk.CTkEntry(frame, width=130, textvariable=self.jq_property_var)
        jq_entry.grid(row=row, column=1, padx=5, pady=3)
        self.jq_property_var.trace_add("write", lambda *args: self._schedule_calc())

    # ── Tab 4: 汇总计算 ──

    def _build_summary_tab(self):
        self.tab_summary.grid_columnconfigure(0, weight=1)
        self.tab_summary.grid_rowconfigure(0, weight=1)

        frame = ctk.CTkScrollableFrame(self.tab_summary, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=16, pady=16)
        frame.grid_columnconfigure(0, weight=1)

        # ── 变量 ──
        self.sum_medical_claim = tk.StringVar(value="0.00")
        self.sum_medical_support = tk.StringVar(value="0.00")
        self.sum_medical_after = tk.StringVar(value="0.00")

        self.sum_death_claim = tk.StringVar(value="0.00")
        self.sum_death_support = tk.StringVar(value="0.00")
        self.sum_death_after = tk.StringVar(value="0.00")

        self.sum_property_claim = tk.StringVar(value="0.00")
        self.sum_property_support = tk.StringVar(value="0.00")
        self.sum_property_after = tk.StringVar(value="0.00")

        self.total_support_var = tk.StringVar(value="0.00")
        self.loss_no_resp_var = tk.StringVar(value="0")
        self.after_jq_total_var = tk.StringVar(value="0.00")
        self.resp_ratio_var = tk.StringVar(value="100")
        self.defendant_pay_var = tk.StringVar(value="0.00")
        self.defendant_advance_var = tk.StringVar(value="0")
        self.actual_pay_var = tk.StringVar(value="0.00")

        # ── Card 1: 三大类赔偿汇总 ──
        card1 = ctk.CTkFrame(frame, fg_color=COLORS["card_bg"], corner_radius=8,
                              border_width=1, border_color=COLORS["card_border"])
        card1.grid(row=0, column=0, sticky="ew", pady=(0, 16))
        for col in range(5):
            card1.grid_columnconfigure(col, weight=0)

        ctk.CTkLabel(card1, text="三大类赔偿汇总",
                      font=(COLORS["font"], 15, "bold"),
                      text_color=COLORS["text_primary"]).grid(
            row=0, column=0, columnspan=5, padx=16, pady=(14, 4), sticky="w")

        headers = ["类别", "原告要求合计", "本院支持合计", "交强险扣除", "扣除后余额"]
        for i, h in enumerate(headers):
            ctk.CTkLabel(card1, text=h, font=(COLORS["font"], 12, "bold"),
                          text_color=COLORS["text_secondary"]).grid(
                row=1, column=i, padx=8, pady=(2, 6))

        categories = [
            ("医疗费类", self.sum_medical_claim, self.sum_medical_support,
             self.jq_medical_var, self.sum_medical_after),
            ("死亡伤残类", self.sum_death_claim, self.sum_death_support,
             self.jq_death_var, self.sum_death_after),
            ("财产类",   self.sum_property_claim, self.sum_property_support,
             self.jq_property_var, self.sum_property_after),
        ]
        for ci, (label, cv, sv, jv, av) in enumerate(categories):
            r = ci + 2
            ctk.CTkLabel(card1, text=label, font=(COLORS["font"], 13),
                          text_color=COLORS["text_primary"], anchor="w").grid(
                row=r, column=0, padx=(14, 6), pady=4, sticky="w")
            for col, var in [(1, cv), (2, sv), (3, jv), (4, av)]:
                ctk.CTkEntry(card1, width=130, height=34, textvariable=var,
                              state="readonly",
                              fg_color="#ffffff", text_color=COLORS["text_primary"],
                              border_width=1, border_color=COLORS["card_border"],
                              corner_radius=6, font=(COLORS["font"], 13)).grid(
                    row=r, column=col, padx=4, pady=3)

        # ── Card 2: 最终计算 ──
        card2 = ctk.CTkFrame(frame, fg_color=COLORS["card_bg"], corner_radius=8,
                              border_width=1, border_color=COLORS["card_border"])
        card2.grid(row=1, column=0, sticky="ew", pady=(0, 16))
        card2.grid_columnconfigure(0, weight=0)
        card2.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(card2, text="最终计算",
                      font=(COLORS["font"], 15, "bold"),
                      text_color=COLORS["text_primary"]).grid(
            row=0, column=0, columnspan=2, padx=16, pady=(14, 6), sticky="w")

        calc_fields = [
            ("法院支持合计:",        self.total_support_var, False, False),
            ("损失不分责:",          self.loss_no_resp_var, True,  False),
            ("扣除交强险后余额:",    self.after_jq_total_var, False, False),
            ("被告承担比例 (%):",    self.resp_ratio_var, True,  False),
            ("被告应赔偿:",          self.defendant_pay_var, False, True),
            ("被告垫付:",            self.defendant_advance_var, True, False),
            ("实际支付原告:",        self.actual_pay_var, False, True),
        ]

        for i, (label, var, editable, highlight) in enumerate(calc_fields):
            r = i + 1
            ctk.CTkLabel(card2, text=label, font=(COLORS["font"], 13,
                          "bold" if highlight else "normal"),
                          text_color=COLORS["text_primary"], anchor="w").grid(
                row=r, column=0, padx=(16, 8), pady=4, sticky="w")

            if highlight:
                entry = ctk.CTkEntry(card2, width=180, height=34,
                                      textvariable=var, state="readonly",
                                      fg_color=COLORS["success_bg"],
                                      text_color=COLORS["accent"],
                                      border_width=0, corner_radius=6,
                                      font=(COLORS["font"], 13, "bold"))
            elif editable:
                entry = ctk.CTkEntry(card2, width=180, height=34,
                                      textvariable=var,
                                      fg_color="#ffffff",
                                      text_color=COLORS["text_primary"],
                                      border_width=1,
                                      border_color=COLORS["card_border"],
                                      corner_radius=6,
                                      font=(COLORS["font"], 13))
            else:
                entry = ctk.CTkEntry(card2, width=180, height=34,
                                      textvariable=var, state="readonly",
                                      fg_color=COLORS["success_bg"],
                                      text_color=COLORS["text_primary"],
                                      border_width=0, corner_radius=6,
                                      font=(COLORS["font"], 13))
            entry.grid(row=r, column=1, padx=8, pady=4, sticky="w")

        # traces for editable fields
        self.loss_no_resp_var.trace_add("write", lambda *args: self._schedule_calc())
        self.resp_ratio_var.trace_add("write", lambda *args: self._schedule_calc())
        self.defendant_advance_var.trace_add("write", lambda *args: self._schedule_calc())

        # ── 计算按钮 ──
        ctk.CTkButton(frame, text="重新计算", command=self._auto_calc,
                       fg_color=COLORS["accent"], hover_color="#1557b0",
                       font=(COLORS["font"], 14, "bold"),
                       height=42, width=200, corner_radius=8).grid(
            row=2, column=0, pady=(0, 16))

    # ── Tab 5: 其他费用 ──

    def _build_other_tab(self):
        self.tab_other.grid_columnconfigure(0, weight=1)
        self.tab_other.grid_rowconfigure(0, weight=1)

        frame = ctk.CTkScrollableFrame(self.tab_other, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=16, pady=16)
        frame.grid_columnconfigure(0, weight=1)

        # ── 变量 ──
        self.appraisal_var = tk.StringVar(value="0")
        self.case_fee_var = tk.StringVar(value="0")
        self.def_case_fee_var = tk.StringVar(value="0")
        self.third_party_var = tk.StringVar(value="0")
        self.total_claim_var = tk.StringVar(value="0.00")

        self.appraisal_var.trace_add("write", lambda *args: self._schedule_calc())
        self.case_fee_var.trace_add("write", lambda *args: self._schedule_calc())
        self.def_case_fee_var.trace_add("write", lambda *args: self._schedule_calc())
        self.third_party_var.trace_add("write", lambda *args: self._schedule_calc())

        # ── Card: 诉讼相关费用 ──
        card = ctk.CTkFrame(frame, fg_color=COLORS["card_bg"], corner_radius=8,
                             border_width=1, border_color=COLORS["card_border"])
        card.grid(row=0, column=0, sticky="ew", pady=(0, 16))
        card.grid_columnconfigure(0, weight=0)
        card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(card, text="诉讼相关费用",
                      font=(COLORS["font"], 15, "bold"),
                      text_color=COLORS["text_primary"]).grid(
            row=0, column=0, columnspan=2, padx=16, pady=(14, 6), sticky="w")

        other_fields = [
            ("鉴定费:",          self.appraisal_var, True),
            ("案件受理费:",      self.case_fee_var, True),
            ("被告承担受理费:",  self.def_case_fee_var, True),
            ("原告三者险应得:",  self.third_party_var, True),
            ("原告诉请合计:",    self.total_claim_var, False),
        ]

        for i, (label, var, editable) in enumerate(other_fields):
            r = i + 1
            is_total = not editable
            ctk.CTkLabel(card, text=label, font=(COLORS["font"], 13,
                          "bold" if is_total else "normal"),
                          text_color=COLORS["text_primary"], anchor="w").grid(
                row=r, column=0, padx=(16, 8), pady=5, sticky="w")

            if editable:
                entry = ctk.CTkEntry(card, width=180, height=34,
                                      textvariable=var,
                                      fg_color="#ffffff",
                                      text_color=COLORS["text_primary"],
                                      border_width=1,
                                      border_color=COLORS["card_border"],
                                      corner_radius=6,
                                      font=(COLORS["font"], 13))
            else:
                entry = ctk.CTkEntry(card, width=180, height=34,
                                      textvariable=var, state="readonly",
                                      fg_color=COLORS["success_bg"],
                                      text_color=COLORS["accent"],
                                      border_width=0, corner_radius=6,
                                      font=(COLORS["font"], 13, "bold"))
            entry.grid(row=r, column=1, padx=8, pady=5, sticky="w")

    # ── Tab 6: 导出 ──

    def _build_export_tab(self):
        self.tab_export.grid_columnconfigure(0, weight=1)
        self.tab_export.grid_rowconfigure(0, weight=1)

        card = ctk.CTkFrame(self.tab_export, fg_color=COLORS["card_bg"], corner_radius=8,
                             border_width=1, border_color=COLORS["card_border"])
        card.grid(row=0, column=0, padx=160, pady=100)

        ctk.CTkLabel(card, text="导出 Excel",
                      font=(COLORS["font"], 18, "bold"),
                      text_color=COLORS["text_primary"]).grid(
            row=0, column=0, pady=(30, 10), padx=40)

        info_text = (
            "将当前所有输入数据按模板格式导出为 Excel 文件。\n\n"
            "导出的文件包含：\n"
            "  · 赔偿汇总 — 三大类赔偿汇总表\n"
            "  · 医疗费类明细 — 医疗费用详细清单\n"
            "  · 死亡伤残类明细 — 死亡伤残费用清单\n"
            "  · 财产类明细 — 财产损失清单"
        )
        ctk.CTkLabel(card, text=info_text, font=(COLORS["font"], 12),
                      text_color=COLORS["text_secondary"],
                      justify="left").grid(row=1, column=0, padx=30, pady=10)

        ctk.CTkButton(
            card, text="选择路径并导出 Excel",
            command=self._export_excel,
            font=(COLORS["font"], 14, "bold"),
            height=46, width=240, corner_radius=8,
            fg_color=COLORS["accent"], hover_color="#1557b0"
        ).grid(row=2, column=0, pady=(20, 30))

    # ══════════════════════════════════════════
    # 计算逻辑
    # ══════════════════════════════════════════

    def _get_float(self, var, default=0.0):
        """从 StringVar 安全获取浮点数"""
        try:
            val = var.get().strip()
            return float(val) if val else default
        except (ValueError, tk.TclError):
            return default

    def _schedule_calc(self, *args):
        """防抖：用户停止输入 300ms 后再计算"""
        if self._calc_after_id:
            self.after_cancel(self._calc_after_id)
        self._calc_after_id = self.after(300, self._auto_calc)

    def _auto_calc(self):
        """自动计算所有汇总和最终金额"""
        try:
            # 1. 读取各项目值
            for item in MEDICAL_ITEMS:
                self.items[item]["claim"] = self._get_float(self.medical_vars[item]["claim"])
                self.items[item]["support"] = self._get_float(self.medical_vars[item]["support"])

            for item in DEATH_DISABILITY_ITEMS:
                self.items[item]["claim"] = self._get_float(self.death_vars[item]["claim"])
                self.items[item]["support"] = self._get_float(self.death_vars[item]["support"])

            for item in PROPERTY_ITEMS:
                self.items[item]["claim"] = self._get_float(self.property_vars[item]["claim"])
                self.items[item]["support"] = self._get_float(self.property_vars[item]["support"])

            # 2. 计算各类小计
            med_claim = sum(self.items[item]["claim"] for item in MEDICAL_ITEMS)
            med_support = sum(self.items[item]["support"] for item in MEDICAL_ITEMS)

            death_claim = sum(self.items[item]["claim"] for item in DEATH_DISABILITY_ITEMS)
            death_support = sum(self.items[item]["support"] for item in DEATH_DISABILITY_ITEMS)

            prop_claim = sum(self.items[item]["claim"] for item in PROPERTY_ITEMS)
            prop_support = sum(self.items[item]["support"] for item in PROPERTY_ITEMS)

            # 3. 更新小计显示
            self.medical_claim_total.set(f"{med_claim:.2f}")
            self.medical_support_total.set(f"{med_support:.2f}")
            self.death_claim_total.set(f"{death_claim:.2f}")
            self.death_support_total.set(f"{death_support:.2f}")
            self.property_claim_total.set(f"{prop_claim:.2f}")
            self.property_support_total.set(f"{prop_support:.2f}")

            # 4. 交强险扣除
            jq_med = self._get_float(self.jq_medical_var)
            jq_death = self._get_float(self.jq_death_var)
            jq_prop = self._get_float(self.jq_property_var)

            # 5. 汇总 Tab 显示
            self.sum_medical_claim.set(f"{med_claim:.2f}")
            self.sum_medical_support.set(f"{med_support:.2f}")
            after_med = max(0, med_support - jq_med)
            self.sum_medical_after.set(f"{after_med:.2f}")

            self.sum_death_claim.set(f"{death_claim:.2f}")
            self.sum_death_support.set(f"{death_support:.2f}")
            after_death = max(0, death_support - jq_death)
            self.sum_death_after.set(f"{after_death:.2f}")

            self.sum_property_claim.set(f"{prop_claim:.2f}")
            self.sum_property_support.set(f"{prop_support:.2f}")
            after_prop = max(0, prop_support - jq_prop)
            self.sum_property_after.set(f"{after_prop:.2f}")

            # 6. 最终计算
            total_support = med_support + death_support + prop_support
            after_jq_total = after_med + after_death + after_prop

            loss_no_resp = self._get_float(self.loss_no_resp_var)
            ratio = self._get_float(self.resp_ratio_var) / 100.0
            defendant_advance = self._get_float(self.defendant_advance_var)

            # 被告应赔偿 = (扣除交强险后余额 + 损失不分责) × 承担比例
            defendant_pay = (after_jq_total + loss_no_resp) * ratio

            # 实际支付 = 被告应赔偿 - 被告垫付
            actual_pay = max(0, defendant_pay - defendant_advance)

            # 其他费用
            appraisal_fee = self._get_float(self.appraisal_var)
            case_fee = self._get_float(self.case_fee_var)
            def_case_fee = self._get_float(self.def_case_fee_var)
            third_party = self._get_float(self.third_party_var)

            total_claim = med_claim + death_claim + prop_claim

            self.total_support_var.set(f"{total_support:.2f}")
            self.after_jq_total_var.set(f"{after_jq_total:.2f}")
            self.defendant_pay_var.set(f"{defendant_pay:.2f}")
            self.actual_pay_var.set(f"{actual_pay:.2f}")
            self.total_claim_var.set(f"{total_claim:.2f}")

            # 存储额外字段供导出使用
            self.extra["损失不分责"] = loss_no_resp
            self.extra["被告支付"] = defendant_pay
            self.extra["总计支持"] = total_support
            self.extra["鉴定费"] = appraisal_fee
            self.extra["案件受理费"] = case_fee
            self.extra["被告承担受理费"] = def_case_fee
            self.extra["被告垫付总额"] = defendant_advance
            self.extra["原告三者险应得"] = third_party
            self.extra["扣除交强险后余额"] = after_jq_total
            self.extra["承担比例"] = ratio
            self.extra["实际支付原告"] = actual_pay

        except Exception:
            pass

    # ══════════════════════════════════════════
    # Excel 导出
    # ══════════════════════════════════════════

    def _export_excel(self):
        """导出 Excel 文件"""
        # 先计算一次确保数据最新
        self._auto_calc()

        # 选择保存路径
        plaintiff = self.plaintiff_name if self.plaintiff_name else "未知原告"
        default_filename = f"{plaintiff}机动车交通事故责任纠纷赔偿计算表.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename,
            title="选择导出路径"
        )
        if not file_path:
            return

        try:
            self._generate_excel(file_path)
            messagebox.showinfo("导出成功", f"文件已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出时发生错误:\n{str(e)}")

    def _generate_excel(self, file_path):
        """生成 Excel 文件"""
        # ── 样式 ──
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        header_font = Font(name='Microsoft YaHei', size=11, bold=True)
        sub_header_font = Font(name='Microsoft YaHei', size=11, bold=True, color="333333")
        data_font = Font(name='Microsoft YaHei', size=11)
        money_fmt = '#,##0.00'

        # 收集数值
        med_claim = sum(self.items[item]["claim"] for item in MEDICAL_ITEMS)
        med_support = sum(self.items[item]["support"] for item in MEDICAL_ITEMS)
        death_claim = sum(self.items[item]["claim"] for item in DEATH_DISABILITY_ITEMS)
        death_support = sum(self.items[item]["support"] for item in DEATH_DISABILITY_ITEMS)
        prop_claim = sum(self.items[item]["claim"] for item in PROPERTY_ITEMS)
        prop_support = sum(self.items[item]["support"] for item in PROPERTY_ITEMS)
        jq_med = self._get_float(self.jq_medical_var)
        jq_death = self._get_float(self.jq_death_var)
        jq_prop = self._get_float(self.jq_property_var)
        total_support_v = self._get_float(self.total_support_var)
        after_jq_v = self._get_float(self.after_jq_total_var)
        defendant_pay_v = self._get_float(self.defendant_pay_var)
        defendant_advance_v = self._get_float(self.defendant_advance_var)
        actual_pay_v = self._get_float(self.actual_pay_var)
        loss_no_resp_v = self._get_float(self.loss_no_resp_var)
        ratio_v = self._get_float(self.resp_ratio_var) / 100.0
        appraisal_fee_v = self._get_float(self.appraisal_var)
        case_fee_v = self._get_float(self.case_fee_var)
        def_case_fee_v = self._get_float(self.def_case_fee_var)
        third_party_v = self._get_float(self.third_party_var)

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════
        # Sheet1: 三大类汇总
        # ════════════════════════════════════════════
        ws = wb.active
        ws.title = "赔偿汇总"

        col_widths = [16, 15, 15, 5, 20, 15, 15, 5, 14, 15, 15, 5, 5, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── 表头行 ──
        headers_top = [
            (1, 1, "医疗费类"), (1, 2, "原告要求"), (1, 3, "本院支持"),
            (1, 5, "死亡伤残赔偿金类"), (1, 6, "原告要求"), (1, 7, "本院支持"),
            (1, 9, "财产类"), (1, 10, "原告要求"), (1, 11, "本院支持"),
            (1, 19, "金额"),
        ]
        for r, c, v in headers_top:
            ws.cell(row=r, column=c, value=v).font = header_font

        # 第一行全部加边框
        for c in range(1, 27):
            ws.cell(row=1, column=c).border = thin_border

        # ── 数据行 ──
        # 格式: (row, col_med, col_death, col_prop)
        rows_layout = [
            (2, "医疗费", "护理费", "财产损失"),
            (3, "住院伙食补助费", "误工费", None),
            (4, "营养费", "残疾赔偿金", None),
            (5, None, "被扶养人生活费", None),
            (6, None, "精神损害抚慰金", None),
            (7, None, "交通费", None),
            (8, None, "护理用品费", None),
        ]

        for r, med_item, death_item, prop_item in rows_layout:
            if med_item:
                ws.cell(row=r, column=1, value=med_item).font = data_font
                ws.cell(row=r, column=2, value=self.items[med_item]["claim"]).font = data_font
                ws.cell(row=r, column=2).number_format = money_fmt
                ws.cell(row=r, column=3, value=self.items[med_item]["support"]).font = data_font
                ws.cell(row=r, column=3).number_format = money_fmt
            if death_item:
                ws.cell(row=r, column=5, value=death_item).font = data_font
                ws.cell(row=r, column=6, value=self.items[death_item]["claim"]).font = data_font
                ws.cell(row=r, column=6).number_format = money_fmt
                ws.cell(row=r, column=7, value=self.items[death_item]["support"]).font = data_font
                ws.cell(row=r, column=7).number_format = money_fmt
            if prop_item:
                ws.cell(row=r, column=9, value=prop_item).font = data_font
                ws.cell(row=r, column=10, value=self.items[prop_item]["claim"]).font = data_font
                ws.cell(row=r, column=10).number_format = money_fmt
                ws.cell(row=r, column=11, value=self.items[prop_item]["support"]).font = data_font
                ws.cell(row=r, column=11).number_format = money_fmt
            for c in range(1, 12):
                ws.cell(row=r, column=c).border = thin_border

        # ── 合计行 ──
        total_r = 10
        ws.cell(row=total_r, column=1, value="合计").font = header_font
        ws.cell(row=total_r, column=2, value=med_claim).font = header_font
        ws.cell(row=total_r, column=2).number_format = money_fmt
        ws.cell(row=total_r, column=3, value=med_support).font = header_font
        ws.cell(row=total_r, column=3).number_format = money_fmt
        ws.cell(row=total_r, column=5, value=death_claim).font = header_font
        ws.cell(row=total_r, column=5).number_format = money_fmt
        ws.cell(row=total_r, column=6, value=death_support).font = header_font
        ws.cell(row=total_r, column=6).number_format = money_fmt
        ws.cell(row=total_r, column=9, value=prop_claim).font = header_font
        ws.cell(row=total_r, column=9).number_format = money_fmt
        ws.cell(row=total_r, column=10, value=prop_support).font = header_font
        ws.cell(row=total_r, column=10).number_format = money_fmt
        for c in range(1, 12):
            ws.cell(row=total_r, column=c).border = thin_border

        # ── 底部计算区域 ──
        calc_start = 12
        calc_items = [
            (12, 10, "损失不分责", loss_no_resp_v),
            (13, 10, "被告支付", defendant_pay_v),
            (14, 10, "总计支持", total_support_v),
            (15, 10, "案件受理费", case_fee_v),
            (16, 10, "被告担", def_case_fee_v),
        ]
        for r, c, lbl, val in calc_items:
            ws.cell(row=r, column=9, value=lbl).font = data_font
            ws.cell(row=r, column=10, value=val).font = data_font
            ws.cell(row=r, column=10).number_format = money_fmt
            ws.cell(row=r, column=9).border = thin_border
            ws.cell(row=r, column=10).border = thin_border

        # 右侧汇总参数
        right_start = 14
        right_headers = [
            ("交强险扣除(医疗)", jq_med),
            ("交强险扣除(伤残)", jq_death),
            ("交强险扣除(财产)", jq_prop),
            ("扣除交强险后余额", after_jq_v),
            ("承担比例", ratio_v),
            ("被告应赔偿", defendant_pay_v),
            ("被告垫付", defendant_advance_v),
            ("实际支付原告", actual_pay_v),
        ]
        for i, (lbl, val) in enumerate(right_headers):
            r = calc_start + i
            ws.cell(row=r, column=14, value=lbl).font = data_font
            ws.cell(row=r, column=15, value=val).font = data_font
            ws.cell(row=r, column=15).number_format = money_fmt if i != 4 else '0%'
            ws.cell(row=r, column=14).border = thin_border
            ws.cell(row=r, column=15).border = thin_border

        # 鉴定费、三者险
        extra_r = calc_start + len(right_headers) + 1
        for i, (lbl, val) in enumerate([("鉴定费", appraisal_fee_v), ("原告三者险应得", third_party_v)]):
            ws.cell(row=extra_r + i, column=14, value=lbl).font = data_font
            ws.cell(row=extra_r + i, column=15, value=val).font = data_font
            ws.cell(row=extra_r + i, column=15).number_format = money_fmt
            ws.cell(row=extra_r + i, column=14).border = thin_border
            ws.cell(row=extra_r + i, column=15).border = thin_border

        # ════════════════════════════════════════════
        # 医疗费类明细
        # ════════════════════════════════════════════
        ws_med = wb.create_sheet(title="医疗费类明细")
        ws_med.column_dimensions['A'].width = 20
        ws_med.column_dimensions['B'].width = 15
        ws_med.column_dimensions['C'].width = 15
        ws_med.column_dimensions['D'].width = 15
        ws_med.column_dimensions['E'].width = 15

        _write_sheet_header(ws_med, ["项目", "原告要求", "本院支持"], header_font, thin_border)
        _write_detail_row(ws_med, 2, "医疗费",
                          self.items["医疗费"]["claim"], self.items["医疗费"]["support"],
                          data_font, header_font, money_fmt, thin_border)

        # 医疗费逐笔明细
        ws_med.cell(row=3, column=1, value="  逐笔明细").font = data_font
        ws_med.cell(row=3, column=1).border = thin_border
        ws_med.cell(row=3, column=2, value="日期").font = sub_header_font
        ws_med.cell(row=3, column=2).border = thin_border
        ws_med.cell(row=3, column=3, value="金额").font = sub_header_font
        ws_med.cell(row=3, column=3).border = thin_border
        med_row = 4
        for date_str, amount in details:
            ws_med.cell(row=med_row, column=2, value=date_str if date_str else "").font = data_font
            ws_med.cell(row=med_row, column=3, value=amount).font = data_font
            ws_med.cell(row=med_row, column=3).number_format = money_fmt
            ws_med.cell(row=med_row, column=2).border = thin_border
            ws_med.cell(row=med_row, column=3).border = thin_border
            med_row += 1

        # 其他医疗项目
        _write_detail_row(ws_med, med_row + 1, "住院伙食补助费",
                          self.items["住院伙食补助费"]["claim"], self.items["住院伙食补助费"]["support"],
                          data_font, header_font, money_fmt, thin_border)
        _write_detail_row(ws_med, med_row + 2, "营养费",
                          self.items["营养费"]["claim"], self.items["营养费"]["support"],
                          data_font, header_font, money_fmt, thin_border)
        # 营养费计算明细
        try:
            nd = float(self.nutrition_days_var.get().strip() or "0")
        except ValueError:
            nd = 0.0
        nutrition_detail_r = med_row + 3
        ws_med.cell(row=nutrition_detail_r, column=1, value=f"  天数: {nd} × 50元/天 = {nd * 50:.2f}元").font = data_font
        ws_med.cell(row=nutrition_detail_r, column=1).border = thin_border
        nutrition_basis_r = med_row + 4
        basis_text = f"  依据: {self.nutrition_basis_var.get()}"
        ws_med.cell(row=nutrition_basis_r, column=1, value=basis_text).font = data_font
        ws_med.cell(row=nutrition_basis_r, column=1).border = thin_border

        # 小计
        sub_r = med_row + 6
        ws_med.cell(row=sub_r, column=1, value="小计").font = header_font
        ws_med.cell(row=sub_r, column=1).border = thin_border
        ws_med.cell(row=sub_r, column=2, value=med_claim).font = header_font
        ws_med.cell(row=sub_r, column=2).number_format = money_fmt
        ws_med.cell(row=sub_r, column=2).border = thin_border
        ws_med.cell(row=sub_r, column=3, value=med_support).font = header_font
        ws_med.cell(row=sub_r, column=3).number_format = money_fmt
        ws_med.cell(row=sub_r, column=3).border = thin_border

        # 交强险
        sub_r += 1
        ws_med.cell(row=sub_r, column=1, value="交强险扣除").font = data_font
        ws_med.cell(row=sub_r, column=1).border = thin_border
        ws_med.cell(row=sub_r, column=3, value=jq_med).font = data_font
        ws_med.cell(row=sub_r, column=3).number_format = money_fmt
        ws_med.cell(row=sub_r, column=3).border = thin_border

        sub_r += 1
        ws_med.cell(row=sub_r, column=1, value="扣除后余额").font = header_font
        ws_med.cell(row=sub_r, column=1).border = thin_border
        ws_med.cell(row=sub_r, column=3, value=max(0, med_support - jq_med)).font = header_font
        ws_med.cell(row=sub_r, column=3).number_format = money_fmt
        ws_med.cell(row=sub_r, column=3).border = thin_border

        # ════════════════════════════════════════════
        # 死亡伤残类明细
        # ════════════════════════════════════════════
        ws_death = wb.create_sheet(title="死亡伤残类明细")
        ws_death.column_dimensions['A'].width = 22
        ws_death.column_dimensions['B'].width = 15
        ws_death.column_dimensions['C'].width = 15
        ws_death.column_dimensions['D'].width = 15
        ws_death.column_dimensions['E'].width = 15

        _write_sheet_header(ws_death, ["项目", "原告要求", "本院支持"], header_font, thin_border)
        for i, item in enumerate(DEATH_DISABILITY_ITEMS):
            _write_detail_row(ws_death, i + 2, item,
                              self.items[item]["claim"], self.items[item]["support"],
                              data_font, header_font, money_fmt, thin_border)

        sub_r = len(DEATH_DISABILITY_ITEMS) + 3
        ws_death.cell(row=sub_r, column=1, value="小计").font = header_font
        ws_death.cell(row=sub_r, column=1).border = thin_border
        ws_death.cell(row=sub_r, column=2, value=death_claim).font = header_font
        ws_death.cell(row=sub_r, column=2).number_format = money_fmt
        ws_death.cell(row=sub_r, column=2).border = thin_border
        ws_death.cell(row=sub_r, column=3, value=death_support).font = header_font
        ws_death.cell(row=sub_r, column=3).number_format = money_fmt
        ws_death.cell(row=sub_r, column=3).border = thin_border

        sub_r += 1
        ws_death.cell(row=sub_r, column=1, value="交强险扣除").font = data_font
        ws_death.cell(row=sub_r, column=1).border = thin_border
        ws_death.cell(row=sub_r, column=3, value=jq_death).font = data_font
        ws_death.cell(row=sub_r, column=3).number_format = money_fmt
        ws_death.cell(row=sub_r, column=3).border = thin_border

        sub_r += 1
        ws_death.cell(row=sub_r, column=1, value="扣除后余额").font = header_font
        ws_death.cell(row=sub_r, column=1).border = thin_border
        ws_death.cell(row=sub_r, column=3, value=max(0, death_support - jq_death)).font = header_font
        ws_death.cell(row=sub_r, column=3).number_format = money_fmt
        ws_death.cell(row=sub_r, column=3).border = thin_border

        # ════════════════════════════════════════════
        # 财产类明细
        # ════════════════════════════════════════════
        ws_prop = wb.create_sheet(title="财产类明细")
        ws_prop.column_dimensions['A'].width = 20
        ws_prop.column_dimensions['B'].width = 15
        ws_prop.column_dimensions['C'].width = 15
        ws_prop.column_dimensions['D'].width = 15
        ws_prop.column_dimensions['E'].width = 15

        _write_sheet_header(ws_prop, ["项目", "原告要求", "本院支持"], header_font, thin_border)
        _write_detail_row(ws_prop, 2, "财产损失",
                          self.items["财产损失"]["claim"], self.items["财产损失"]["support"],
                          data_font, header_font, money_fmt, thin_border)

        sub_r = 4
        ws_prop.cell(row=sub_r, column=1, value="小计").font = header_font
        ws_prop.cell(row=sub_r, column=1).border = thin_border
        ws_prop.cell(row=sub_r, column=2, value=prop_claim).font = header_font
        ws_prop.cell(row=sub_r, column=2).number_format = money_fmt
        ws_prop.cell(row=sub_r, column=2).border = thin_border
        ws_prop.cell(row=sub_r, column=3, value=prop_support).font = header_font
        ws_prop.cell(row=sub_r, column=3).number_format = money_fmt
        ws_prop.cell(row=sub_r, column=3).border = thin_border

        sub_r += 1
        ws_prop.cell(row=sub_r, column=1, value="交强险扣除").font = data_font
        ws_prop.cell(row=sub_r, column=1).border = thin_border
        ws_prop.cell(row=sub_r, column=3, value=jq_prop).font = data_font
        ws_prop.cell(row=sub_r, column=3).number_format = money_fmt
        ws_prop.cell(row=sub_r, column=3).border = thin_border

        sub_r += 1
        ws_prop.cell(row=sub_r, column=1, value="扣除后余额").font = header_font
        ws_prop.cell(row=sub_r, column=1).border = thin_border
        ws_prop.cell(row=sub_r, column=3, value=max(0, prop_support - jq_prop)).font = header_font
        ws_prop.cell(row=sub_r, column=3).number_format = money_fmt
        ws_prop.cell(row=sub_r, column=3).border = thin_border

        # 保存
        wb.save(file_path)


def _write_sheet_header(ws, headers, font, border):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font
        cell.border = border


def _write_detail_row(ws, row, label, claim_val, support_val, font, bold_font, money_fmt, border):
    ws.cell(row=row, column=1, value=label).font = font
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=claim_val).font = font
    ws.cell(row=row, column=2).number_format = money_fmt
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=3, value=support_val).font = font
    ws.cell(row=row, column=3).number_format = money_fmt
    ws.cell(row=row, column=3).border = border


# ══════════════════════════════════════════
# 入口
# ══════════════════════════════════════════

if __name__ == "__main__":
    app = CompensationApp()
    app.mainloop()
